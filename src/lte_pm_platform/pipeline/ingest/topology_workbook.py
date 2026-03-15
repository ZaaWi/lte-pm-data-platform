from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ParsedTopologySnapshotRow:
    source_row_number: int
    logical_entity_key: str
    dataset_family: str
    site_code: str
    site_name: str
    region_code: str
    region_name: str
    area_name: str | None
    cluster_id: str | None
    team_code: str | None
    reporting_key: str | None
    reporting_name: str | None
    reporting_level: str | None
    workbook_subnet_id: str
    workbook_enodeb_id: str
    workbook_enodeb_name: str
    workbook_cell_name: str
    mapping_source: str
    notes: str | None


@dataclass(frozen=True)
class ParsedTopologyWorkbook:
    source_file_name: str
    topology_release_date: date | None
    source_sha256: str
    workbook_row_count: int
    normalized_rows: list[ParsedTopologySnapshotRow]
    parser_warnings: list[str]
    parser_errors: list[str]


_RELEASE_DATE_PATTERN = re.compile(r"(\d{8})")
_REQUIRED_COLUMNS = {
    "SubnetID",
    "eNodeBid",
    "ENODEBName",
    "CellID",
    "CELLNAME",
    "SiteName",
    "Region",
}


def extract_release_date_from_filename(file_name: str) -> date | None:
    match = _RELEASE_DATE_PATTERN.search(file_name)
    if not match:
        return None
    return datetime.strptime(match.group(1), "%Y%m%d").date()


def sha256_file(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_topology_workbook(file_path: Path) -> ParsedTopologyWorkbook:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if "4G LTE" not in workbook.sheetnames:
        raise ValueError("Workbook is missing required sheet: 4G LTE")

    worksheet = workbook["4G LTE"]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError("Workbook sheet 4G LTE is empty")

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    missing_columns = sorted(column for column in _REQUIRED_COLUMNS if column not in headers)
    if missing_columns:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing_columns)}")

    index = {header: position for position, header in enumerate(headers)}
    workbook_row_count = 0
    normalized_rows: list[ParsedTopologySnapshotRow] = []
    parser_warnings: list[str] = []
    parser_errors: list[str] = []

    for source_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        workbook_row_count += 1
        row_data = {header: _normalize_cell(row[position]) for header, position in index.items()}
        try:
            normalized_rows.extend(_normalize_workbook_row(row_data, source_row_number))
        except ValueError as exc:
            parser_errors.append(f"row {source_row_number}: {exc}")

    return ParsedTopologyWorkbook(
        source_file_name=file_path.name,
        topology_release_date=extract_release_date_from_filename(file_path.name),
        source_sha256=sha256_file(file_path),
        workbook_row_count=workbook_row_count,
        normalized_rows=normalized_rows,
        parser_warnings=parser_warnings,
        parser_errors=parser_errors,
    )


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_workbook_row(row: dict[str, str], source_row_number: int) -> list[ParsedTopologySnapshotRow]:
    subnet_id = _required_value(row, "SubnetID")
    enodeb_id = _required_value(row, "eNodeBid")
    enodeb_name = _required_value(row, "ENODEBName")
    cell_id = _required_value(row, "CellID")
    cell_name = _required_value(row, "CELLNAME")
    site_name = _required_value(row, "SiteName")
    region_name = _required_value(row, "Region")

    site_code = site_name
    region_code = region_name
    area_name = row.get("Area") or None
    cluster_id = row.get("ClusterID") or None
    team_code = row.get("TEAM") or None
    reporting_key = f"CLUSTER::{cluster_id}" if cluster_id else None
    reporting_name = cluster_id if cluster_id else None

    common = {
        "source_row_number": source_row_number,
        "site_code": site_code,
        "site_name": site_name,
        "region_code": region_code,
        "region_name": region_name,
        "area_name": area_name,
        "cluster_id": cluster_id,
        "team_code": team_code,
        "reporting_key": reporting_key,
        "reporting_name": reporting_name,
        "reporting_level": "cluster" if cluster_id else None,
        "workbook_subnet_id": subnet_id,
        "workbook_enodeb_id": enodeb_id,
        "workbook_enodeb_name": enodeb_name,
        "workbook_cell_name": cell_name,
        "mapping_source": "project_parameter_workbook",
        "notes": None,
    }

    return [
        ParsedTopologySnapshotRow(
            logical_entity_key=f"family=PM/sdr/ltefdd|sbnid={subnet_id}|enodebid={enodeb_id}|cellid={cell_id}",
            dataset_family="PM/sdr/ltefdd",
            **common,
        ),
        ParsedTopologySnapshotRow(
            logical_entity_key=f"family=PM/itbbu/ltefdd|sbnid={subnet_id}|enbid={enodeb_id}|cellid={cell_id}",
            dataset_family="PM/itbbu/ltefdd",
            **common,
        ),
    ]


def _required_value(row: dict[str, str], key: str) -> str:
    value = (row.get(key) or "").strip()
    if not value:
        raise ValueError(f"missing critical field: {key}")
    return value
